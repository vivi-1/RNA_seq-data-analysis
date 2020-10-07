#!/usr/bin/env python3
# -*- coding: utf-8 -*-
##################Developed by Wei Wang and Kevin Yu##########################
import os
import pandas
import re
import xlrd
import pandas
import re
import openpyxl
import xlwt
from openpyxl import Workbook
import xlrd

# from xlrd import load_workbook
currentDirectory = "/Volumes/WD1/DEGlist/Comparison files/comparison to other studies"
os.getcwd()
os.chdir(currentDirectory)
Name_newfile=[]
## Give new file name here

#data downloaded from: https://link.springer.com/article/10.1007/s10725-016-0163-1#Sec17
Name_newfile='Nbenthi under P.parasitica.xlsx'
highlight_workbook=xlrd.open_workbook(Name_newfile,'rb')
highlight_sheet=highlight_workbook.sheet_by_index(0)
highlight_ID=[]
for i in range(0,highlight_sheet.nrows):
    for j in range(0,highlight_sheet.ncols): 
        j=0 ## Read the 1st column
    highlight_ID.append(highlight_sheet.cell(i,j).value)
highlight_ID=highlight_ID[3:]

## NB0.4.4 transfered to NB1.0.1
df = pandas.read_csv('Niben044vsNiben101_GeneID_mapping.txt', sep='\t',names=['NbC','Niben101Scf'])
df=df.stack().str.replace(r'.', '').unstack()
df['NbC']=df['NbC'].str.slice(0, -1)
df['Niben101Scf']=df['Niben101Scf'].str.slice(0, -1)
list_temp=df['NbC'].values.tolist()
wb=[]
ws=[]
wb=xlwt.Workbook('Nbenthi under P.parasitica_NBC to NB101.xls','wrb')
ws=wb.add_sheet('sheet1')
for i in range(0,len(highlight_ID)):
    ID_value=highlight_ID[i]
    if ID_value in list_temp:
         gene_to_write_row_number=df.NbC[df.NbC == ID_value].index.tolist()
         gene_to_write_row_number=int(str(gene_to_write_row_number).strip('[]'))
         ws.write(i+1,0,ID_value)
         value_to_write=df.loc[gene_to_write_row_number].at['Niben101Scf']
         ws.write(i+1,1,value_to_write)
         print(i,'Nb101 found')
    else:
        ws.write(i+1,0,ID_value)
        ws.write(i+1,1,'no corresponding Nb101 found')
        print(i,'Nb101 not found')
    
ws.write(0,0,'NbC')
ws.write(0,1,'Niben101Scf')
wb.save('Nbenthi under P.parasitica_NBC to NB101.xls')

################################################
#data downloaded from: https://link.springer.com/article/10.1007/s10725-016-0163-1#Sec17
group_name='group7'
newfile_dir="/Volumes/WD1/DEGlist/Comparison files/Gene_ID for each group/"
Name_newfile=newfile_dir+'timepoints combined '+group_name+' with annotation.xls'
regex=re.compile(r'\d+')
k=int(regex.search(group_name).group(0))
sheetname1=str(k)
read_sheet=pandas.read_excel(Name_newfile, index_col=None, usecols="A", sheet_name=sheetname1)
read_sheet_list=read_sheet.values.tolist()
#search files  
searchfile_name='Nbenthi under P.parasitica_NBC to NB101.xls'
searchfile_dir="/Volumes/WD1/DEGlist/Comparison files/comparison to other studies/"
Name_searchfile=searchfile_dir+searchfile_name
Name_searchfile_workbook=xlrd.open_workbook(Name_searchfile,'rb')
search_sheet=Name_searchfile_workbook.sheet_by_index(0)
search_ID=[]
for i in range(0,search_sheet.nrows):
    for j in range(0,search_sheet.ncols): 
        j=1 ## Read the second column
    search_ID.append(search_sheet.cell(i,j).value)
search_ID=search_ID[1:]
# highlight_ID = [x for x in highlight_ID if x != 'no corresponding Nb101 found']

## Get number as k from the Name_groupfile

wb=[]
ws1=[]
wb=xlwt.Workbook('top200_comparison with P.parasitica_Group'+sheetname1+'.xls','wrb')
ws1=wb.add_sheet(sheetname1)
print("sheet"+ sheetname1 + "is created")
for i in range(0,len(read_sheet_list)):
    ID_value=[read_sheet_list[i]]
    ws1.write(i+1,0,read_sheet_list[i])
    if ID_value in search_ID:
        ws1.write(i+1,1,'Y')
        #ws1.write(i,j,value_to_write)    
    else:
        ws1.write(i+1,1,'N')
ws1.write(0,0,'Gene_ID')
ws1.write(0,1,'If the gene is in P.parasitica file')
wb.save('top200_comparison with P.parasitica_Group'+sheetname1+'.xls')
print("data is written in "+"sheet"+sheetname1)



        
    
    
    
    