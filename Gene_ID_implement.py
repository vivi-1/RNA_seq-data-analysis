#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun May 16 00:09:32 2021

@author: weiwang
"""

import openpyxl as xl;
import csv
import pandas as pd
from pandas import DataFrame

def gene_ID_obtain(source, filename):
    print("Start")    
    wb1 = xl.load_workbook(source)
    ws1 = wb1.worksheets[0]
    new_rows_list = []
    print(ws1.cell(74,1).value)
    
    print("Read Gene_ID file")
    fileIn = open(filename, 'r') 
    reader = csv.reader(fileIn)
    r = next(reader)
    for row in reader:
        print(row[1])
        print(ws1.cell(int(row[1]) + 1, 1).value)
        new_rows_list.append(ws1.cell(int(row[1]) + 1, 1).value)
    fileIn.close() 
    #print(new_rows_list)
    
    df1 = DataFrame(new_rows_list,columns=['Gene_ID'])
    print("Write to file")
    df1.to_csv(filename, mode='a', index=False)
for i in range(1, 8):
    filename = "/Users/weiwang/Desktop/DEGlist/Output/Eliminate outlier rep2/T" + str(i) + "_down_Flag22_Pnic_no rep2.csv"
    source = "/Users/weiwang/Desktop/DEGlist/Input/Gene_ID_list.xlsx"
    gene_ID_obtain(source, filename)
for i in range(1, 8):
    filename = "/Users/weiwang/Desktop/DEGlist/Output/Eliminate outlier rep2/T" + str(i) + "_up_Flag22_Pnic_no rep2.csv"
    source = "/Users/weiwang/Desktop/DEGlist/Input/Gene_ID_list.xlsx"
    gene_ID_obtain(source, filename)
for i in range(1, 8):
    filename = "/Users/weiwang/Desktop/DEGlist/Output/Eliminate outlier rep2/T" + str(i) + "_down_Pnic_no rep2.csv"
    source = "/Users/weiwang/Desktop/DEGlist/Input/Gene_ID_list.xlsx"
    gene_ID_obtain(source, filename)
for i in range(1, 8):
    filename = "/Users/weiwang/Desktop/DEGlist/Output/Eliminate outlier rep2/T" + str(i) + "_up_Pnic_no rep2.csv"
    source = "/Users/weiwang/Desktop/DEGlist/Input/Gene_ID_list.xlsx"
    gene_ID_obtain(source, filename)
for i in range(1, 8):
    filename = "/Users/weiwang/Desktop/DEGlist/Output/Eliminate outlier rep2/T" + str(i) + "_down_Flag22_no rep2.csv"
    source = "/Users/weiwang/Desktop/DEGlist/Input/Gene_ID_list.xlsx"
    gene_ID_obtain(source, filename)
for i in range(1, 8):
    filename = "/Users/weiwang/Desktop/DEGlist/Output/Eliminate outlier rep2/T" + str(i) + "_up_Flag22_no rep2.csv"
    source = "/Users/weiwang/Desktop/DEGlist/Input/Gene_ID_list.xlsx"
    gene_ID_obtain(source, filename)
 
